# Guilherme Dambroski

# Inserção de valores com conversão e taxa INGRAM - AWS

from openpyxl import load_workbook
import pandas as pd
import sys

file = r'C:\Users\TI\OneDrive - IG Guilherme TI\OneDrive - Irmaos Goncalves Com Ind Ltda\GIT\fatura_aws\Custo_AWS.xlsx'

dt_uso = input('Insira a o mês e ano respectivos à cobrança (mm/yyyy): ')
print(dt_uso)
print()

dt_emissao_nf = input('Insira a data da emissão da DANFE (dd/mm/yyyy): ')
print(dt_emissao_nf)
print()

n_nf = input('Insira o Nº da DANFE: ')
print(n_nf)
print()

print('Para os valores solicitados à seguir, não utilize separadores de milhar e separe as casas decimais com "." (ponto)')
print()

vlr_nf = float(input('Insira o valor que consta na DANFE R$: '))
print(vlr_nf)
print()

aws_igerp = float(input('Insira o valor que consta no Console AWS-IGERP(fora marketplace) $: '))
print(aws_igerp)
print()

aws_igerp_marketplace = float(input('Insira o valor que consta no Console AWS-IGERP(somente marketplace) $: '))
print(aws_igerp)
print()

aws_bi = float(input('Insira o valor que consta no Console AWS-BI $: '))
print(aws_bi)
print()

aws_infor = float(input('Insira o valor que consta no Console AWS-INFOR $: '))
print(aws_infor)
print()

cambio = float(input('Insira o valor do câmbio ($ -> R$) que consta na DANFE: '))
print(cambio)
print()

taxa_aws = float('1.1383')
taxa_marketplace = float('1.3256')
cambio_marketplace = cambio * 0.9749582975534787 #***

# Valor por conta AWS x câmbio (USD -> BRL) x taxa fixa cobrada pela AWS para receber o valor total usado, sem imposto 
vlr_final_igerp = aws_igerp * cambio * taxa_aws
vlr_final_igerp_marketplace = aws_igerp_marketplace * cambio * taxa_marketplace
vlr_final_igerp = vlr_final_igerp + vlr_final_igerp_marketplace
vlr_final_bi = aws_bi * cambio * taxa_aws
vlr_final_infor = aws_infor * cambio * taxa_aws

# Soma das contas (BRL)
vlr_final_total = vlr_final_igerp + vlr_final_bi + vlr_final_infor

# Soma das contas (USD)
vlr_final_dol = aws_igerp + aws_bi + aws_infor

# Diferença de valor total entre a DANFE e o painel de faturamento AWS
nf_vlr_dif = vlr_nf - vlr_final_total

print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
print()

print(f'{"Emissão DANFE":.<30}[{dt_emissao_nf}]')
print(f'{"Cobrança":.<30}[{dt_uso}]')
print(f'{"Imposto":.<30}[{taxa_aws}]')
print(f'{"Câmbio":.<30}[{cambio}]')
print()
print(f'{"361659537551-IGERP":.<30}[R${vlr_final_igerp:.2f}]')
print(f'{"164884010347-BI":.<30}[R${vlr_final_bi:.2f}]')
print(f'{"909198593940-INFOR":.<30}[R${vlr_final_infor:.2f}]')
print()
print(f'{"Valor DANFE":.<30}[R${vlr_nf:.2f}]')
print(f'{"Valor AWS":.<30}[R${vlr_final_total:.2f}]')
print(f'{"Diferença":.<30}[R${nf_vlr_dif:.2f}]')
print()

print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
print()

obs = input('Caso tenha alguma observação para adicionar, por favor digite-a. Caso não, apenas pressione "enter". ')
print()

save_excel = input('Solicitação para inserção dos dados (Salvamento no Excel) - ("s" para "Sim" e "n" para "Não"). ')
print()


if save_excel == 'n':
  sys.exit()

elif save_excel == 's':

  if obs == '':

    df_file_old = pd.read_excel(file)

    df = pd.DataFrame({
      'Mês cobrado':   [dt_uso           ,                   '',                '',                  '', '##########'],
      'Emissão DANFE': [dt_emissao_nf    ,                   '',                '',                  '', '##########'],
      'Nº DANFE':      [n_nf             ,                   '',                '',                  '', '##########'],
      'Conta':         ['308486287573-TI', '361659537551-IGERP', '164884010347-BI', '909198593940-INFOR', '##########'],
      'AWS $':         [vlr_final_dol    ,            aws_igerp,            aws_bi,           aws_infor, '##########'],
      'AWS R$':        [vlr_final_total  ,      vlr_final_igerp,      vlr_final_bi,     vlr_final_infor, '##########'],
      'Taxa':          [taxa_aws         ,                   '',                '',                  '', '##########'],
      'Câmbio':        [cambio           ,                   '',                '',                  '', '##########'],
      'DANFE R$':      [vlr_nf           ,                   '',                '',                  '', '##########'],
      'Total R$':      [vlr_final_total  ,                   '',                '',                  '', '##########'],
      'Diferença R$':  [nf_vlr_dif       ,                   '',                '',                  '', '##########'],
      'Observação':    ['***'            ,                   '',                '',                  '', '##########']
    })
    df_concat = pd.concat([df_file_old, df], ignore_index=True)

    #carrego o Excel com o template pré-formatado
    book = load_workbook(file)

    #defino o writer para escrever em um novo arquivo
    writer = pd.ExcelWriter(file, engine='xlsxwriter')

    #incluo a formatação no writer
    # writer.book = df_file_old
    # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df_concat.to_excel(writer, sheet_name='Custo AWS', index=False)

    writer.save()
    
    print()
    print('Arquivo Custo_AWS.xlsx atualizado!')
    print(f'Salvo em {file}')
    print()

  else:

    df_file_old = pd.read_excel(file)

    df = pd.DataFrame({
      'Mês cobrado':   [dt_uso           ,                   '',                '',                  '', '##########'],
      'Emissão DANFE': [dt_emissao_nf    ,                   '',                '',                  '', '##########'],
      'Nº DANFE':      [n_nf             ,                   '',                '',                  '', '##########'],
      'Conta':         ['308486287573-TI', '361659537551-IGERP', '164884010347-BI', '909198593940-INFOR', '##########'],
      'AWS $':         [vlr_final_dol    ,            aws_igerp,            aws_bi,           aws_infor, '##########'],
      'AWS R$':        [vlr_final_total  ,      vlr_final_igerp,      vlr_final_bi,     vlr_final_infor, '##########'],
      'Taxa':          [taxa_aws         ,                   '',                '',                  '', '##########'],
      'Câmbio':        [cambio           ,                   '',                '',                  '', '##########'],
      'DANFE R$':      [vlr_nf           ,                   '',                '',                  '', '##########'],
      'Total R$':      [vlr_final_total  ,                   '',                '',                  '', '##########'],
      'Diferença R$':  [nf_vlr_dif       ,                   '',                '',                  '', '##########'],
      'Observação':    [obs              ,                   '',                '',                  '', '##########']
    })
    df_concat = pd.concat([df_file_old, df], ignore_index=True)

    #carrego o Excel com o template pré-formatado
    book = load_workbook(file)

    #defino o writer para escrever em um novo arquivo
    writer = pd.ExcelWriter(file, engine='xlsxwriter')

    #incluo a formatação no writer
    # writer.book = df_file_old
    # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df_concat.to_excel(writer, sheet_name='Custo AWS', index=False)

    writer.save()
    
    print()
    print('Arquivo Custo_AWS.xlsx atualizado!')
    print(f'Salvo em {file}')
    print()

else:
  print('Opção inválida. Saindo! ')
  sys.exit()